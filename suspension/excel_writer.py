# suspension/excel_writer.py
# Project-GP — Alex Suspension Excel Writer
# =============================================================================
#
# Generates Front_Ter27_Alex.xlsx and Rear_Ter27_Alex.xlsx from the optimizer
# output. Maintains the exact Optimum Kinematics format of the Velis files
# (all hardpoint rows unchanged) and:
#   - Updates: Static Camber, Static Toe, Spring, U-Bar rows
#   - Adds a new sheet "Alex Optimization" with:
#       · Full heave sweep table (camber, toe, MR, RC, scrub vs z)
#       · Pareto front summary table
#       · Steer sweep table (front only)
#       · Roll analysis table
#       · Kinematic gains summary
#
# SHEET LAYOUT:
#   Sheet 1: "Front/Rear Suspension"  — identical structure to Velis
#             Hardpoints: UNCHANGED (frozen)
#             Stiffness rows: UPDATED with Alex values
#             Wheel rows:     UPDATED with Alex camber/toe
#   Sheet 2: "Alex Optimization"      — NEW
#             § A: Kinematic Gains Summary
#             § B: Heave Sweep (500 rows)
#             § C: Steer Sweep (100 rows, front only)
#             § D: Roll Analysis (100 rows)
#             § E: Pareto Front (from MORL)
#             § F: Recommended Setups (Skidpad, Endurance, Balanced)
# =============================================================================

from __future__ import annotations

import math
import datetime
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple

import numpy as np

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers as xl_numbers,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_00
except ImportError:
    raise ImportError("openpyxl required: pip install openpyxl --break-system-packages")

from suspension.sweep_analysis import SweepResult, SteerSweepResult, RollAnalysisResult


# ---------------------------------------------------------------------------
# §1  Style constants
# ---------------------------------------------------------------------------

_ACCENT    = "1A3C6E"   # dark racing blue
_RED       = "C0392B"
_LIGHT     = "D6E4F0"
_MIDBLUE   = "AEC6E8"
_GREY      = "F5F5F5"
_WHITE     = "FFFFFF"
_DARKTEXT  = "1C1C1C"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=10, color=_DARKTEXT, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _header_font() -> Font:
    return Font(bold=True, size=10, color=_WHITE, name="Calibri")


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center")


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# §2  Cell helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, row: int, col: int, labels: List[str]) -> None:
    """Write a row of header cells with accent background."""
    for j, label in enumerate(labels, start=col):
        c = ws.cell(row=row, column=j, value=label)
        c.font      = _header_font()
        c.fill      = _fill(_ACCENT)
        c.alignment = _center()
        c.border    = _thin_border()


def _write_data_row(
    ws, row: int, col: int,
    values: List[Any],
    fill: Optional[str] = None,
    bold: bool = False,
    fmt: Optional[str] = None,
) -> None:
    for j, v in enumerate(values, start=col):
        c = ws.cell(row=row, column=j, value=v)
        c.font      = _font(bold=bold)
        c.alignment = _center()
        c.border    = _thin_border()
        if fill:
            c.fill = _fill(fill)
        if fmt and isinstance(v, (int, float)):
            c.number_format = fmt


def _set_col_widths(ws, widths: Dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _section_title(ws, row: int, col: int, text: str, n_cols: int) -> None:
    """Merged section title cell."""
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=col + n_cols - 1
    )
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=12, color=_WHITE, name="Calibri")
    c.fill      = _fill(_ACCENT)
    c.alignment = _left()
    c.border    = _thin_border()


# ---------------------------------------------------------------------------
# §3  Update Velis sheet with Alex parameters
# ---------------------------------------------------------------------------

def _update_velis_sheet(
    ws,
    alex_camber_deg: float,
    alex_toe_deg:    float,
    alex_spring_nmm: float,   # N/mm  (Optimum K format)
    alex_arb_nmrad:  float,   # N·m/rad → N·m/deg for display
    analysis_date:   datetime.date,
) -> None:
    """
    Scan the existing Velis sheet and update ONLY the mutable cells:
        - 'Static Camber' row, column C and G  (left/right)
        - 'Static Toe'    row, column C and G
        - 'Spring'        row, column D and H  (N/mm)
        - 'U-Bar'         row, column F         (N·m/deg)
        - 'Name'          row → update to Alex
        - 'Version'       row → update date
    """
    alex_arb_nm_per_deg = alex_arb_nmrad * (math.pi / 180.0)

    for row in ws.iter_rows():
        label_cell = row[1] if len(row) > 1 else None
        if label_cell is None or label_cell.value is None:
            continue
        label = str(label_cell.value).strip()

        if label == "Static Camber":
            if len(row) > 2 and row[2].value is not None:
                row[2].value = round(alex_camber_deg, 4)
            if len(row) > 6 and row[6].value is not None:
                row[6].value = round(alex_camber_deg, 4)

        elif label == "Static Toe":
            if len(row) > 2 and row[2].value is not None:
                row[2].value = round(alex_toe_deg, 4)
            if len(row) > 6 and row[6].value is not None:
                row[6].value = round(alex_toe_deg, 4)

        elif label == "Spring":
            # Column D (index 3) = left; column H (index 7) = right
            if len(row) > 3 and row[3].value is not None:
                row[3].value = round(alex_spring_nmm, 2)
            if len(row) > 7 and row[7].value is not None:
                row[7].value = round(alex_spring_nmm, 2)

        elif label == "U-Bar":
            # Column F (index 5) = middle (symmetric)
            if len(row) > 5 and row[5].value is not None:
                row[5].value = round(alex_arb_nm_per_deg, 4)

    # Update Name row (column C = index 2)
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and "Velis" in cell.value:
                cell.value = cell.value.replace("Velis", "Alex")
                cell.value = cell.value.replace("Velis 2", "Alex")

    # Update Modified on / Version
    for row in ws.iter_rows(max_row=5):
        label = str(row[1].value).strip() if row[1].value else ""
        if label == "Version:":
            # Row has: (None, 'Version:', ver, 'Modified by:', author, None, 'Modified on:', date)
            if len(row) > 7:
                row[7].value = datetime.datetime.now()


# ---------------------------------------------------------------------------
# §4  Alex Optimization sheet builder
# ---------------------------------------------------------------------------

def _build_alex_sheet(
    ws,
    sweep_f:   SweepResult,
    sweep_r:   SweepResult,
    steer:     Optional[SteerSweepResult],
    roll_f:    RollAnalysisResult,
    roll_r:    RollAnalysisResult,
    pareto_setups:   Optional[np.ndarray],   # (N_pareto, 28) or None
    pareto_grips:    Optional[np.ndarray],
    pareto_stabs:    Optional[np.ndarray],
    setup_names:     Optional[List[str]],
    recommended:     Optional[Dict[str, Dict[str, float]]],
) -> None:
    """
    Build the 'Alex Optimization' sheet with 6 sections.
    All sections are placed vertically in a single sheet, separated by blank rows.
    """
    row_cursor = [1]   # mutable so sub-functions can advance it

    def next_row(n=1):
        r = row_cursor[0]
        row_cursor[0] += n
        return r

    def blank(n=1):
        row_cursor[0] += n

    # ── §A: Kinematic Gains Summary ───────────────────────────────────────────
    r = next_row()
    _section_title(ws, r, 1, "§ A — Kinematic Gains Summary", 8)
    blank()

    r = next_row()
    _write_header_row(ws, r, 1, [
        "Axle", "Camber Gain [deg/m]", "Bump Steer [deg/m]",
        "MR @ z=0", "RC Height @ z=0 [mm]", "dRC/dz [-]",
        "KPI [deg]", "Caster [deg]"
    ])
    for sweep in [sweep_f, sweep_r]:
        r = next_row()
        _write_data_row(ws, r, 1, [
            sweep.axle.upper(),
            round(sweep.camber_gain_deg_per_m, 3),
            round(sweep.bump_steer_deg_per_m * 1000.0, 3),  # deg/m → mdeg/m for readability
            round(sweep.mr_at_zero, 4),
            round(sweep.rc_at_zero_mm, 2),
            round(sweep.drc_dz, 4),
            round(sweep.kpi_static_deg, 2),
            round(sweep.caster_static_deg, 2),
        ], fill=_LIGHT, fmt="0.000")

    blank(2)

    # ── §B: Heave Sweep Table ─────────────────────────────────────────────────
    r = next_row()
    _section_title(ws, r, 1, "§ B — Heave Sweep (Front: cols 1–9, Rear: cols 11–19)", 19)
    blank()

    sweep_headers = [
        "z [mm]", "Camber [deg]", "Toe [deg]", "Caster [deg]",
        "KPI [deg]", "MR [-]", "RC [mm]", "Scrub [mm]", "Track Δ [mm]"
    ]
    r = next_row()
    _write_header_row(ws, r, 1,  ["FRONT"] + sweep_headers[1:])
    _write_header_row(ws, r, 11, ["REAR"]  + sweep_headers[1:])

    # Use the same z-grid; if front and rear have different n_pts just use front
    n = len(sweep_f.z_mm)
    for i in range(min(n, len(sweep_r.z_mm))):
        r = next_row()
        fill = _GREY if i % 2 == 0 else _WHITE
        _write_data_row(ws, r, 1, [
            round(float(sweep_f.z_mm[i]),          2),
            round(float(sweep_f.camber_deg[i]),     4),
            round(float(sweep_f.toe_deg[i]),        5),
            round(float(sweep_f.caster_deg[i]),     4),
            round(float(sweep_f.kpi_deg[i]),        4),
            round(float(sweep_f.motion_ratio[i]),   5),
            round(float(sweep_f.rc_height_mm[i]),   2),
            round(float(sweep_f.scrub_radius_mm[i]),2),
            round(float(sweep_f.track_change_mm[i]),2),
        ], fill=fill, fmt="0.0000")
        _write_data_row(ws, r, 11, [
            round(float(sweep_r.z_mm[i]),          2),
            round(float(sweep_r.camber_deg[i]),     4),
            round(float(sweep_r.toe_deg[i]),        5),
            round(float(sweep_r.caster_deg[i]),     4),
            round(float(sweep_r.kpi_deg[i]),        4),
            round(float(sweep_r.motion_ratio[i]),   5),
            round(float(sweep_r.rc_height_mm[i]),   2),
            round(float(sweep_r.scrub_radius_mm[i]),2),
            round(float(sweep_r.track_change_mm[i]),2),
        ], fill=fill, fmt="0.0000")

    blank(2)

    # ── §C: Steer Sweep (front only) ──────────────────────────────────────────
    r = next_row()
    if steer is not None:
        _section_title(ws, r, 1, "§ C — Steer Sweep (Front Axle, Ackermann Analysis)", 6)
        blank()
        r = next_row()
        _write_header_row(ws, r, 1, [
            "Rack Travel [mm]", "Outer Toe [deg]", "Inner Toe [deg]",
            "Ackermann %", "Ideal Ackermann [deg]", "Delta Toe [deg]"
        ])
        for i in range(len(steer.rack_travel_mm)):
            r = next_row()
            fill = _GREY if i % 2 == 0 else _WHITE
            _write_data_row(ws, r, 1, [
                round(float(steer.rack_travel_mm[i]),  2),
                round(float(steer.toe_outer_deg[i]),   4),
                round(float(steer.toe_inner_deg[i]),   4),
                round(float(steer.ackermann_pct[i]),   1),
                round(float(steer.ideal_ackermann[i]), 4),
                round(float(steer.toe_inner_deg[i] - steer.toe_outer_deg[i]), 4),
            ], fill=fill, fmt="0.000")
    else:
        _section_title(ws, r, 1, "§ C — Steer Sweep (N/A: rear passive steer axle)", 6)

    blank(2)

    # ── §D: Roll Analysis ─────────────────────────────────────────────────────
    r = next_row()
    _section_title(ws, r, 1,  "§ D — Roll Analysis (Front: cols 1–7, Rear: cols 9–15)", 15)
    blank()

    roll_headers = [
        "Roll [deg]", "Camber Outer [deg]", "Camber Inner [deg]",
        "ΔCamber [deg]", "Toe Outer [deg]", "Toe Inner [deg]", "RC [mm]"
    ]
    r = next_row()
    _write_header_row(ws, r, 1, roll_headers)
    _write_header_row(ws, r, 9, roll_headers)

    for i in range(len(roll_f.roll_deg)):
        r = next_row()
        fill = _GREY if i % 2 == 0 else _WHITE
        _write_data_row(ws, r, 1, [
            round(float(roll_f.roll_deg[i]),          3),
            round(float(roll_f.camber_outer_deg[i]),  4),
            round(float(roll_f.camber_inner_deg[i]),  4),
            round(float(roll_f.camber_net_deg[i]),    4),
            round(float(roll_f.toe_outer_deg[i]),     4),
            round(float(roll_f.toe_inner_deg[i]),     4),
            round(float(roll_f.rc_height_mm[i]),      2),
        ], fill=fill, fmt="0.0000")
        _write_data_row(ws, r, 9, [
            round(float(roll_r.roll_deg[i]),          3),
            round(float(roll_r.camber_outer_deg[i]),  4),
            round(float(roll_r.camber_inner_deg[i]),  4),
            round(float(roll_r.camber_net_deg[i]),    4),
            round(float(roll_r.toe_outer_deg[i]),     4),
            round(float(roll_r.toe_inner_deg[i]),     4),
            round(float(roll_r.rc_height_mm[i]),      2),
        ], fill=fill, fmt="0.0000")

    blank(2)

    # ── §E: Pareto Front ──────────────────────────────────────────────────────
    r = next_row()
    if pareto_setups is not None and setup_names is not None:
        n_pareto = pareto_setups.shape[0]
        n_cols   = min(pareto_setups.shape[1] + 2, 30)
        _section_title(ws, r, 1, f"§ E — MORL Pareto Front ({n_pareto} setups)", n_cols)
        blank()

        r = next_row()
        pareto_headers = ["Grip [G]", "Stability [rad/s]"] + list(setup_names)
        _write_header_row(ws, r, 1, pareto_headers[:n_cols])

        for i in range(n_pareto):
            r = next_row()
            fill = _GREY if i % 2 == 0 else _WHITE
            row_vals = (
                [round(float(pareto_grips[i]), 4),
                 round(float(pareto_stabs[i]), 4)]
                + [round(float(v), 5) for v in pareto_setups[i]]
            )
            _write_data_row(ws, r, 1, row_vals[:n_cols], fill=fill, fmt="0.0000")
    else:
        _section_title(ws, r, 1,
                       "§ E — MORL Pareto Front (run MORL optimizer to populate)", 10)

    blank(2)

    # ── §F: Recommended Setups ────────────────────────────────────────────────
    r = next_row()
    _section_title(ws, r, 1, "§ F — Recommended Setups", 10)
    blank()

    if recommended:
        r = next_row()
        _write_header_row(ws, r, 1, [
            "Event", "Toe F [deg]", "Camber F [deg]",
            "Spring F [N/mm]", "ARB F [N·m/deg]",
            "Toe R [deg]", "Camber R [deg]",
            "Spring R [N/mm]", "ARB R [N·m/deg]",
            "Notes"
        ])
        for event, params in recommended.items():
            r = next_row()
            _write_data_row(ws, r, 1, [
                event,
                round(params.get("toe_f_deg", 0.0), 3),
                round(params.get("camber_f_deg", -0.5), 3),
                round(params.get("spring_f_nmm", 44.0), 1),
                round(params.get("arb_f_nmdeg", 5.0), 2),
                round(params.get("toe_r_deg", 0.0), 3),
                round(params.get("camber_r_deg", -1.0), 3),
                round(params.get("spring_r_nmm", 53.0), 1),
                round(params.get("arb_r_nmdeg", 5.0), 2),
                params.get("notes", ""),
            ], fill=_LIGHT, fmt="0.000")
    else:
        r = next_row()
        ws.cell(row=r, column=1, value="Run MORL optimizer to generate recommendations.")

    # ── Column widths ─────────────────────────────────────────────────────────
    _set_col_widths(ws, {
        1: 18, 2: 16, 3: 16, 4: 14, 5: 14,
        6: 14, 7: 14, 8: 14, 9: 14, 10: 20,
        11: 18, 12: 16, 13: 16, 14: 14, 15: 14,
        16: 14, 17: 14, 18: 14, 19: 14,
    })


# ---------------------------------------------------------------------------
# §5  Main writer function
# ---------------------------------------------------------------------------

def write_alex_suspension_excel(
    front_template_path: str,
    rear_template_path:  str,
    output_dir:          str,
    sweep_f:             SweepResult,
    sweep_r:             SweepResult,
    roll_f:              RollAnalysisResult,
    roll_r:              RollAnalysisResult,
    # Optimized setup parameters (from MORL output)
    alex_front: Dict[str, float],
    alex_rear:  Dict[str, float],
    # Optional MORL Pareto front data
    pareto_setups:  Optional[np.ndarray] = None,
    pareto_grips:   Optional[np.ndarray] = None,
    pareto_stabs:   Optional[np.ndarray] = None,
    setup_names:    Optional[List[str]]  = None,
    steer_sweep:    Optional[SteerSweepResult] = None,
    recommended:    Optional[Dict[str, Dict[str, float]]] = None,
) -> Tuple[str, str]:
    """
    Generate Front_Ter27_Alex.xlsx and Rear_Ter27_Alex.xlsx.

    Args:
        front_template_path : path to Front_Ter27_-_Velis.xlsx
        rear_template_path  : path to Rear_TeR27_-_Velis_2.xlsx
        output_dir          : directory to write Alex files into
        sweep_f, sweep_r    : SweepResult from compute_sweep()
        roll_f, roll_r      : RollAnalysisResult from compute_roll_analysis()
        alex_front          : dict with keys:
                                camber_deg, toe_deg, spring_nmm, arb_nmrad
        alex_rear           : same keys
        pareto_*            : optional MORL results
        steer_sweep         : optional front steer sweep
        recommended         : dict of event name → parameter dict
                                e.g. {'Skidpad': {'toe_f_deg': -0.1, ...}}

    Returns:
        (front_path, rear_path) : absolute paths of generated files.
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.date.today()

    for (template, axle, suffix, sweep, roll, params) in [
        (front_template_path, "front", "Front_Ter27_Alex.xlsx",
         sweep_f, roll_f, alex_front),
        (rear_template_path,  "rear",  "Rear_Ter27_Alex.xlsx",
         sweep_r, roll_r, alex_rear),
    ]:
        print(f"[ExcelWriter] Writing {suffix}...")
        wb = load_workbook(template)
        ws_susp = wb.active

        # ── 1. Update existing Velis sheet ────────────────────────────────────
        _update_velis_sheet(
            ws_susp,
            alex_camber_deg = params["camber_deg"],
            alex_toe_deg    = params["toe_deg"],
            alex_spring_nmm = params["spring_nmm"],
            alex_arb_nmrad  = params["arb_nmrad"],
            analysis_date   = now,
        )

        # ── 2. Add Alex Optimization sheet ────────────────────────────────────
        ws_alex = wb.create_sheet("Alex Optimization")

        _build_alex_sheet(
            ws_alex,
            sweep_f=sweep_f if axle == "front" else sweep_r,
            sweep_r=sweep_r if axle == "front" else sweep_f,
            steer=steer_sweep if axle == "front" else None,
            roll_f=roll_f,
            roll_r=roll_r,
            pareto_setups=pareto_setups,
            pareto_grips=pareto_grips,
            pareto_stabs=pareto_stabs,
            setup_names=setup_names,
            recommended=recommended,
        )

        # ── 3. Save ────────────────────────────────────────────────────────────
        out_path = out_dir / suffix
        wb.save(str(out_path))
        print(f"[ExcelWriter] Saved → {out_path}")

    front_path = str(out_dir / "Front_Ter27_Alex.xlsx")
    rear_path  = str(out_dir / "Rear_Ter27_Alex.xlsx")
    return front_path, rear_path


# ---------------------------------------------------------------------------
# §6  Convenience: generate Alex files from a complete analysis run
# ---------------------------------------------------------------------------

def generate_alex_files_from_analysis(
    front_template_path: str,
    rear_template_path:  str,
    output_dir:          str,
    front_kin,           # SuspensionKinematics
    rear_kin,            # SuspensionKinematics
    front_kin_right,     # SuspensionKinematics (right side, for steer sweep)
    rear_kin_right,      # SuspensionKinematics or None
    # Target setup parameters (from MORL or manual)
    toe_f_deg:        float = -0.10,
    camber_f_deg:     float = -2.50,
    spring_f_nmm:     float = 42.0,
    arb_f_nmrad:      float = 600.0,
    toe_r_deg:        float =  0.15,
    camber_r_deg:     float = -2.20,
    spring_r_nmm:     float = 56.0,
    arb_r_nmrad:      float = 400.0,
    vp:               Optional[Dict[str, Any]] = None,
    # Optional MORL results
    pareto_setups:    Optional[np.ndarray] = None,
    pareto_grips:     Optional[np.ndarray] = None,
    pareto_stabs:     Optional[np.ndarray] = None,
    setup_names:      Optional[List[str]]  = None,
) -> Tuple[str, str]:
    """
    One-call convenience wrapper: runs all analyses then writes both Excel files.

    Typical usage:
        from suspension.excel_writer import generate_alex_files_from_analysis
        front_path, rear_path = generate_alex_files_from_analysis(
            'data/Front_Ter27_-_Velis.xlsx',
            'data/Rear_TeR27_-_Velis_2.xlsx',
            'output/',
            front_kin, rear_kin, front_kin_right, rear_kin_right,
            toe_f_deg=-0.10, camber_f_deg=-2.5, ...
        )
    """
    from suspension.sweep_analysis import (
        compute_sweep, compute_steer_sweep, compute_roll_analysis
    )

    vp = vp or {}

    print("[generate_alex_files] Running kinematic analyses...")
    sweep_f = compute_sweep(front_kin, toe_f_deg, camber_f_deg, axle="front")
    sweep_r = compute_sweep(rear_kin,  toe_r_deg, camber_r_deg, axle="rear")

    steer = compute_steer_sweep(
        front_kin, front_kin_right,
        toe_f_deg, camber_f_deg,
        axle="front",
        wheelbase_m=vp.get("lf", 0.8525) + vp.get("lr", 0.6975),
    )

    roll_f = compute_roll_analysis(
        front_kin, front_kin_right,
        toe_f_deg, camber_f_deg,
        track_m=vp.get("track_front", 2.0 * front_kin.half_track),
    )
    roll_r = compute_roll_analysis(
        rear_kin, rear_kin_right or rear_kin,
        toe_r_deg, camber_r_deg,
        track_m=vp.get("track_rear", 2.0 * rear_kin.half_track),
    )

    recommended = {
        "Skidpad (Alex)": {
            "toe_f_deg": toe_f_deg, "camber_f_deg": camber_f_deg,
            "spring_f_nmm": spring_f_nmm, "arb_f_nmdeg": arb_f_nmrad * math.pi / 180.0,
            "toe_r_deg": toe_r_deg, "camber_r_deg": camber_r_deg,
            "spring_r_nmm": spring_r_nmm, "arb_r_nmdeg": arb_r_nmrad * math.pi / 180.0,
            "notes": "MORL Pareto grip-maximizing setup",
        },
        "Velis (reference)": {
            "toe_f_deg": 0.0, "camber_f_deg": -0.5,
            "spring_f_nmm": 44.0, "arb_f_nmdeg": 5.0,
            "toe_r_deg": 0.0, "camber_r_deg": -1.0,
            "spring_r_nmm": 53.0, "arb_r_nmdeg": 5.0,
            "notes": "Velis baseline (hardpoints unchanged)",
        },
    }

    front_path, rear_path = write_alex_suspension_excel(
        front_template_path=front_template_path,
        rear_template_path=rear_template_path,
        output_dir=output_dir,
        sweep_f=sweep_f,
        sweep_r=sweep_r,
        roll_f=roll_f,
        roll_r=roll_r,
        alex_front={
            "camber_deg": camber_f_deg,
            "toe_deg":    toe_f_deg,
            "spring_nmm": spring_f_nmm,
            "arb_nmrad":  arb_f_nmrad,
        },
        alex_rear={
            "camber_deg": camber_r_deg,
            "toe_deg":    toe_r_deg,
            "spring_nmm": spring_r_nmm,
            "arb_nmrad":  arb_r_nmrad,
        },
        pareto_setups=pareto_setups,
        pareto_grips=pareto_grips,
        pareto_stabs=pareto_stabs,
        setup_names=setup_names,
        steer_sweep=steer,
        recommended=recommended,
    )

    return front_path, rear_path