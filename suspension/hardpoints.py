# suspension/hardpoints.py
# Project-GP — Velis Hardpoint Parser
# =============================================================================
#
# Parses the Optimum Kinematics-format Excel files produced for the Ter27
# suspension and returns typed, unit-converted dictionaries suitable for
# SuspensionKinematics.
#
# Units contract (internal, after this module):
#   Lengths  : METRES  (Excel is millimetres → divide by 1000)
#   Angles   : DEGREES (Excel is degrees, converted to radians only at
#               the point of consumption in kinematics.py)
#   Forces   : N/m  (spring N/mm → multiply by 1000)
#   Stiffness: N·m/rad (ARB N·m/deg → multiply by 180/π)
#
# Coordinate system (Optimum K body-fixed):
#   X  : longitudinal, forward positive
#   Y  : lateral, LEFT positive   (so the left wheel has Y > 0)
#   Z  : vertical, UP positive
#
# All hardpoint dicts contain NUMPY arrays (not JAX arrays).
# SuspensionKinematics converts them to JAX at __init__ time.
# =============================================================================

from __future__ import annotations

import math
import numpy as np
from pathlib import Path
from typing import Dict, Any

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError(
        "openpyxl is required: pip install openpyxl --break-system-packages"
    )


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _mm(v: float) -> float:
    """Convert mm → m."""
    return v / 1000.0


def _point(x, y, z) -> np.ndarray:
    """Return a (3,) float64 array in metres."""
    return np.array([_mm(x), _mm(y), _mm(z)], dtype=np.float64)


def _nm_per_deg_to_rad(v: float) -> float:
    """Convert N·m/deg → N·m/rad."""
    return v * (180.0 / math.pi)


def _n_per_mm_to_m(v: float) -> float:
    """Convert N/mm → N/m."""
    return v * 1000.0


# ---------------------------------------------------------------------------
# Excel row scanner
# ---------------------------------------------------------------------------

def _scan_wb(path: Path) -> Dict[str, Any]:
    """
    Load the first sheet of an Optimum Kinematics .xlsx export and return a
    flat dict of {point_name: np.ndarray([X, Y, Z])} for all hardpoints,
    plus scalar entries for stiffness and wheel parameters.

    The parser is deliberately lenient about row ordering; it finds named rows
    by matching cell text, not by hard-coded row indices.
    """
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Collect all rows as lists of (possibly-None) values
    rows = [list(row) for row in ws.iter_rows(values_only=True)]

    result: Dict[str, Any] = {}

    # ── Hardpoint extraction ─────────────────────────────────────────────────
    # In the Optimum K format, hardpoints appear as:
    #   (None, 'PointName', X_left, Y_left, Z_left, None, X_right, Y_right, Z_right)
    # The section header rows (e.g. 'Double A-Arm') mark the start of each group.
    # We scan every row; if column B is a non-empty string and columns C/D/E are
    # numeric, we treat it as a hardpoint row.
    #
    # For symmetric (left/right) hardpoints:
    #   Y_left  > 0  (left side)
    #   Y_right < 0  (right side = mirror image)
    # We always store the LEFT side in the dict.

    for row in rows:
        if len(row) < 5:
            continue

        name_cell = row[1]
        if not isinstance(name_cell, str) or not name_cell.strip():
            continue

        name = name_cell.strip()

        # --- Numeric triplet in columns C, D, E ---
        try:
            x = float(row[2])
            y = float(row[3])
            z = float(row[4])
        except (TypeError, ValueError):
            # Not a point row — try scalar fields below
            pass
        else:
            result[name] = _point(x, y, z)  # LEFT side, metres
            # Also store right side if present (cols G, H, I = indices 6, 7, 8)
            if len(row) >= 9:
                try:
                    xr = float(row[6])
                    yr = float(row[7])
                    zr = float(row[8])
                    result[name + "_R"] = _point(xr, yr, zr)
                except (TypeError, ValueError):
                    pass
            continue

        # --- Scalar fields ---
        # "Half Track", "Static Camber", "Static Toe", etc.
        # Value in column C (index 2)
        if len(row) >= 3:
            try:
                scalar_val = float(row[2])
                result[name] = scalar_val
            except (TypeError, ValueError):
                pass

    # ── Stiffness extraction ─────────────────────────────────────────────────
    # The 'Stiffness' section has rows like:
    #   (None, 'Spring',  'Push Pull', left_val, None, None, None, right_val, None)
    #   (None, 'U-Bar',   'U-Bar',     None,     None, mid_val, None, None,   None)

    for row in rows:
        if len(row) < 6:
            continue
        label = row[1] if isinstance(row[1], str) else ""
        label = label.strip()

        if label == "Spring":
            # Left spring: col D (index 3), units: N/mm → N/m
            try:
                result["spring_rate_N_per_m"] = _n_per_mm_to_m(float(row[3]))
            except (TypeError, ValueError):
                pass

        if label == "U-Bar":
            # Middle U-bar stiffness: col F (index 5), units: N·m/deg → N·m/rad
            try:
                result["ubar_stiffness_Nm_per_rad"] = _nm_per_deg_to_rad(float(row[5]))
            except (TypeError, ValueError):
                pass

    # ── Steering ratio ───────────────────────────────────────────────────────
    for row in rows:
        if len(row) < 3:
            continue
        label = row[1] if isinstance(row[1], str) else ""
        if "Steering Ratio" in label:
            try:
                result["steering_ratio_mm_per_rev"] = float(row[2])
            except (TypeError, ValueError):
                pass

    wb.close()
    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_front_hardpoints(xlsx_path: str | Path) -> Dict[str, Any]:
    """
    Parse Front_Ter27_Velis.xlsx (or any front Optimum K file).

    Returns a dict with keys (all LEFT side, metric):
      Double A-Arm:
        'CHAS_LowFor', 'CHAS_LowAft'              — lower A-arm chassis pivots
        'CHAS_UppFor', 'CHAS_UppAft'              — upper A-arm chassis pivots
        'UPRI_LowPnt'                              — lower ball joint
        'UPRI_UppPnt'                              — upper ball joint
        'CHAS_TiePnt', 'UPRI_TiePnt'              — tie-rod chassis / upright
      Push-Pull (PUSHROD — attaches to lower A-arm):
        'NSMA_PPAttPnt_L'                          — pushrod upright attachment
        'CHAS_AttPnt_L'                            — pushrod chassis end
        'CHAS_RocAxi_L', 'CHAS_RocPiv_L'          — rocker axis (2-point definition)
        'ROCK_RodPnt_L'                            — rocker-rod attachment
        'ROCK_CoiPnt_L'                            — rocker-coilover attachment
      U-Bar:
        'NSMA_UBarAttPnt_L'                        — ARB attachment on rocker
        'UBAR_AttPnt_L'                            — ARB blade outer point
        'CHAS_PivPnt_L'                            — ARB chassis pivot
      Scalars (converted):
        'Half Track'               [m]
        'Static Camber'            [deg]
        'Static Toe'               [deg]
        'Rim Diameter'             [mm — kept as mm for reference]
        'Tire Diameter'            [mm — kept as mm for reference]
        'spring_rate_N_per_m'      [N/m]
        'ubar_stiffness_Nm_per_rad'[N·m/rad]
        'steering_ratio_mm_per_rev'[mm/rev]
      Derived:
        'R_wheel'          [m]  — (Tire Diameter / 2) / 1000
        'actuation_type'   str  — 'pushrod'
    """
    raw = _scan_wb(Path(xlsx_path))

    # Derive wheel radius
    if "Tire Diameter" in raw:
        raw["R_wheel"] = _mm(float(raw["Tire Diameter"]) / 2.0)

    # Convert Half Track
    if "Half Track" in raw:
        raw["Half Track_m"] = _mm(float(raw["Half Track"]))

    raw["actuation_type"] = "pushrod"   # confirmed from Attachment section
    raw["_source_file"] = str(xlsx_path)
    return raw


def load_rear_hardpoints(xlsx_path: str | Path) -> Dict[str, Any]:
    """
    Parse Rear_TeR27_Velis_2.xlsx (or any rear Optimum K file).

    Identical structure to front, except:
      - Pullrod attaches to the UPPER A-arm (not lower)
      - Tierod attachment: Chassis (passive rear steer — no rack input)
      - 'actuation_type' == 'pullrod'

    The kinematic solver treats pushrod/pullrod via 'actuation_type' flag,
    which determines which A-arm the rod-to-upright attachment point is on.
    """
    raw = _scan_wb(Path(xlsx_path))

    if "Tire Diameter" in raw:
        raw["R_wheel"] = _mm(float(raw["Tire Diameter"]) / 2.0)

    if "Half Track" in raw:
        raw["Half Track_m"] = _mm(float(raw["Half Track"]))

    raw["actuation_type"] = "pullrod"   # confirmed: NSMA_PPAttPnt attaches to UpperAArm
    raw["passive_rear_steer"] = True    # no rack input; tierod is chassis-fixed
    raw["_source_file"] = str(xlsx_path)
    return raw


def validate_hardpoints(hpts: Dict[str, Any], axle: str = "front") -> None:
    """
    Assert that all mandatory hardpoints are present and internally consistent.
    Raises KeyError on missing point, ValueError on geometric violation.
    """
    mandatory = [
        "CHAS_LowFor", "CHAS_LowAft",
        "CHAS_UppFor", "CHAS_UppAft",
        "UPRI_LowPnt", "UPRI_UppPnt",
        "CHAS_TiePnt", "UPRI_TiePnt",
        "NSMA_PPAttPnt_L", "CHAS_AttPnt_L",
        "CHAS_RocAxi_L", "CHAS_RocPiv_L",
        "ROCK_RodPnt_L", "ROCK_CoiPnt_L",
        "NSMA_UBarAttPnt_L", "UBAR_AttPnt_L", "CHAS_PivPnt_L",
    ]
    for k in mandatory:
        if k not in hpts:
            raise KeyError(f"[hardpoints/{axle}] Missing mandatory hardpoint: '{k}'")
        arr = hpts[k]
        if not isinstance(arr, np.ndarray) or arr.shape != (3,):
            raise ValueError(f"[hardpoints/{axle}] '{k}' must be a (3,) ndarray, got {type(arr)}")

    # Geometric sanity: upper ball joint must be above lower ball joint
    dz = hpts["UPRI_UppPnt"][2] - hpts["UPRI_LowPnt"][2]
    if dz < 0.05:   # 50 mm minimum separation
        raise ValueError(
            f"[hardpoints/{axle}] Upper ball joint Z ({hpts['UPRI_UppPnt'][2]*1e3:.1f} mm) "
            f"must be at least 50 mm above lower ball joint Z "
            f"({hpts['UPRI_LowPnt'][2]*1e3:.1f} mm). Got dz = {dz*1e3:.1f} mm."
        )

    # Geometric sanity: tie rod must have non-trivial length
    tie_len = np.linalg.norm(hpts["UPRI_TiePnt"] - hpts["CHAS_TiePnt"])
    if tie_len < 0.10:   # 100 mm
        raise ValueError(
            f"[hardpoints/{axle}] Tie-rod length ({tie_len*1e3:.1f} mm) < 100 mm — "
            f"check UPRI_TiePnt / CHAS_TiePnt coordinates."
        )

    print(f"[hardpoints/{axle}] Validation passed. "
          f"KP span = {dz*1e3:.1f} mm, tie-rod = {tie_len*1e3:.1f} mm.")


# ---------------------------------------------------------------------------
# Debug helper
# ---------------------------------------------------------------------------

def print_hardpoints(hpts: Dict[str, Any]) -> None:
    """Pretty-print all hardpoints (mm) and scalars."""
    print(f"\n[hardpoints] Source: {hpts.get('_source_file', '?')}")
    print(f"  actuation_type : {hpts.get('actuation_type', '?')}")
    print(f"  R_wheel        : {hpts.get('R_wheel', float('nan'))*1e3:.2f} mm")
    print(f"  Half Track     : {hpts.get('Half Track_m', float('nan'))*1e3:.2f} mm")
    print(f"  spring_rate    : {hpts.get('spring_rate_N_per_m', float('nan'))/1000:.1f} N/mm")
    print(f"  ARB stiffness  : {hpts.get('ubar_stiffness_Nm_per_rad', float('nan')):.2f} N·m/rad")
    print()
    for k, v in sorted(hpts.items()):
        if isinstance(v, np.ndarray) and v.shape == (3,):
            print(f"  {k:<26}  [{v[0]*1e3:+8.3f},  {v[1]*1e3:+8.3f},  {v[2]*1e3:+8.3f}]  mm")